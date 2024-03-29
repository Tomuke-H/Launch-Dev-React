import uuid
import requests
from flask import Flask, jsonify, request, render_template, session, url_for, redirect, send_from_directory, make_response, Response
from flask_session import Session # https://pythonhosted.org/Flask-Session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.sql import text
import msal 
import msal  
import struct
import pyodbc
import os
from os.path import join, dirname, realpath
import pandas as pd
import io
import flask_excel as excel
from pyexcel_xls import get_data
from sqlalchemy.sql.expression import insert
from sql import getSQLConnection
from datetime import date, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook







# when working local, set Local to True and copy app_config to app_config_local to put in values.  This will be in Git ignore and won't be pulled into source.  
Local = False

if Local is False:
    import app_config as app_config
else:
    import app_config_local as app_config


#import pandas as pd

app = Flask(__name__)
app.config.from_object(app_config)
#sess = Session() # the standard session
#sess.init_app(app)
#Session(app)


# This section is needed for url_for("foo", _external=True) to automatically
# generate http scheme when this sample is running on localhost,
# and to generate https scheme when it is deployed behind reversed proxy.
# See also https://flask.palletsprojects.com/en/1.0.x/deploying/wsgi-standalone/#proxy-setups
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)



#Routing
@app.route('/')
def home():
    #un comment below for auth.
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    return render_template('index.html')

@app.route("/login")
def login():
    # Technically we could use empty list [] as scopes to do just sign in,
    # here we choose to also collect end user consent upfront
    session["flow"] = _build_auth_code_flow(scopes=app_config.SCOPE)
    return render_template("login.html", auth_url=session["flow"]["auth_uri"], version=msal.__version__)

@app.route('/launchplans')
def launchesearch():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    return render_template('launchplans.html')   

@app.route('/launchinsights')
def launchinsights():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    return render_template('launchinsights.html')

@app.route('/launchprofile')
def launches():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    return render_template('launchprofile.html')

@app.route('/launchmapping')
def launchmapping():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    return render_template('launchmapping.html')

@app.route('/freightcalculator')
def freightcalculator():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    return render_template('freightcalculator.html')
@app.route('/masterdata')
def masterdata():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    return render_template('masterdata.html')

#@app.route('/visits-counter/')
#def visits():
    #for i in session:
    #    print(i, session[i])
    #if 'visits' in session:
    #    session['visits'] = session.get('visits') + 1  # reading and updating session data
    #else:
    #    session['visits'] = 1 # setting session data
    #return "Total visits: {}".format(session.get('visits'))

@app.route(app_config.REDIRECT_PATH)  # Its absolute URL must match your app's redirect_uri set in AAD
def authorized():
    try:
        cache = _load_cache()
        result = _build_msal_app(cache=cache).acquire_token_by_auth_code_flow(
            session.get("flow", {}), request.args)
        if "error" in result:
            return render_template("auth_error.html", result=result)
        session["user"] = result.get("id_token_claims")
        _save_cache(cache)
    except ValueError:  # Usually caused by CSRF
        pass  # Simply ignore them
    return redirect(url_for("/"))

@app.route("/logout")
def logout():
    session.clear()  # Wipe out user and its token cache from session
    return redirect(  # Also logout from your tenant's web session
        app_config.AUTHORITY + "/oauth2/v2.0/logout" +
        "?post_logout_redirect_uri=" + url_for("/", _external=True))

# Auth Helper Functions




def _load_cache():
    cache = msal.SerializableTokenCache()
    if session.get("token_cache"):
        cache.deserialize(session["token_cache"])
    return cache
def _save_cache(cache):
    if cache.has_state_changed:
        session["token_cache"] = cache.serialize()
def _build_msal_app(cache=None, authority=None):
    return msal.ConfidentialClientApplication(
        app_config.CLIENT_ID, authority=authority or app_config.AUTHORITY,
        client_credential=app_config.CLIENT_SECRET, token_cache=cache)
def _build_auth_code_flow(authority=None, scopes=None):
    return _build_msal_app(authority=authority).initiate_auth_code_flow(
        scopes or [],
        redirect_uri=url_for("authorized", _external=True))
def _get_token_from_cache(scope=None):
    cache = _load_cache()  # This web app maintains one cache per session
    cca = _build_msal_app(cache=cache)
    accounts = cca.get_accounts()
    if accounts:  # So all account(s) belong to the current signed-in user
        result = cca.acquire_token_silent(scope, account=accounts[0])
        _save_cache(cache)
        return result

# SQL Helper functions

def _getSQLToken():
    clientSecret = app_config.CLIENT_SECRET
    clientID = app_config.CLIENT_ID
    authority_url = app_config.AUTHORITY
    context = msal.ConfidentialClientApplication(client_id=clientID, client_credential=clientSecret,authority=authority_url)
    token = context.acquire_token_for_client(scopes=[app_config.SQLRESOURCE])
    return token

#APIS     
#/launchprofiles/<string:name>/

#Not getting Notes, Desc, or Launch Type


@app.route('/launchprofiles', methods=['GET', 'POST'])
def launchprofiles():
    launchprofileparameter = request.args.get('launchprofileparameter')
    xstr = lambda s: None if s == '' else s
    #print(launchprofileparameter)
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method == 'POST':
        data = request.form.to_dict()
        print(data)
        #df = pd.DataFrame.from_dict(data)
        #return(print(region))
        conn = getSQLConnection(app_config=app_config)
        with conn.cursor() as cursor:
            insert = text('INSERT INTO [launchmodeldev].[dbo].[FactLaunchProfiles] VALUES(NEWID(),?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)')
            name  = data.get('launchprofilenameid')
            LOB = data.get('lobid')
            codeName = data.get('codenameid')
            existingSKUProfile = data.get('skuprofileid')
            Description = data.get('descriptionid')
            POMPOD = data.get('pompoddid')
            LaunchDate = data.get('launchdateid')
            LaunchType = data.get('launchtypeid')
            Regions = data.get('regionstring')
            AnnounceDate = data.get('announcedateid') #I have to fix the name here and in UI
            AnnounceFlag = data.get('announcedid')
            AOCIPQ = data.get('aocipq')
            EOCIPQ = data.get('eocipq')
            APOCIPQ = data.get('apocipq')
            LOCIPQ = data.get('locipq')
            FCCDate = data.get('fccdateid')
            PQSDate = data.get('pqsdateid')
            DCVolume = data.get('dcvolume')
            DTSVolume = data.get('dtsvolume')
            MSStoreIPQ = data.get("msstoreipq")
            Notes = data.get('notesid')
            ChangeDate = datetime.now() 
            Createdby = 'chosbo@microsoft.com'
            params =(name,LOB,codeName,existingSKUProfile,Description,POMPOD,LaunchDate,LaunchType,Regions,xstr(AnnounceDate),AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,xstr(FCCDate),xstr(PQSDate),DCVolume,DTSVolume,MSStoreIPQ,Notes,ChangeDate,Createdby)
            print(params)
            cursor.execute(str(insert),params)
            print("Entered")

            

        return 'Success'
        #redirect(url_for("launchprofile"))
        #'Success' redirect(url_for("login")) return redirect(url_for("/LaunchPlans"))
        #     
    if request.method == "GET":
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        results = []
        wb = Workbook()
        my_sheet = wb.active
        #wb.title = "Launch-Profile-Template"
        print(launchprofileparameter)
        print("Workbook Created")
        if launchprofileparameter == "All" :
            wb.title = "All-Launch Profile(s)"
        else :
            wb.title = launchprofileparameter + "- Launch Profile(s)"
        print(wb.title)
        #my_sheet = wb.create_sheet("Sheet_A")
        #my_sheet.title = "Launch-Profile-Template"
        #ws2 = wb.create_sheet("Sheet_B", 0)
        #ws2.title = "Title_B"

        my_sheet['A1'].value = "Launch-Profile-Template"
        my_sheet['A3'].value = "Required"
        my_sheet['B3'].value = "Required"
        my_sheet['C3'].value = "Required"
        my_sheet['D3'].value = "Required"
        my_sheet['F3'].value = "Required"
        my_sheet['G3'].value = "Required"
        my_sheet['H3'].value = "Required"
        my_sheet['I3'].value = "Required"
        my_sheet['K3'].value = "Required"
        my_sheet['A4'].value = "LaunchProfileName"
        my_sheet['B4'].value = "LineOfBusiness"
        my_sheet['C4'].value = "CodeName"
        my_sheet['D4'].value = "ExistingSKUProfile"
        my_sheet['E4'].value = "Description"
        my_sheet['F4'].value = "POM/POD"
        my_sheet['G4'].value = "LaunchDate"
        my_sheet['H4'].value = "LaunchType"
        my_sheet['I4'].value = "Region(s)"
        my_sheet['J4'].value = "AnnounceDate"
        my_sheet['K4'].value = "Announced(Y/N)"
        my_sheet['L4'].value = "AOCIPQ"
        my_sheet['M4'].value = "EOCIPQ"
        my_sheet['N4'].value = "APOCIPQ"
        my_sheet['O4'].value = "LOCIPQ"
        my_sheet['P4'].value = "FCCDate"
        my_sheet['Q4'].value = "PQSDate"
        my_sheet['R4'].value = "DCVolume"
        my_sheet['S4'].value = "DTSVolume"
        my_sheet['T4'].value = "MSStoreIPQ"
        my_sheet['U4'].value = "Notes"
        #my_sheet.title = "LaunchProfileTemplate"

        with conn.cursor() as cursor:
            selectall = text('SELECT DISTINCT Name as LaunchProfileName, LOB as LineOfBusiness,CodeName,ExistingSKUProfile,Description,POMPOD as [POM/POD],CONVERT(varchar,LaunchDate,101) as LaunchDate,LaunchType,Regions as [Region(s)],CONVERT(varchar,AnnounceDate,101) as AnnounceDate,AnnounceFlag as [Announced(Y/N)],AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,CONVERT(varchar,PQSDate,101) as PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes FROM [launchmodeldev].[dbo].[FactLaunchProfiles]')
            selectone = text('SELECT DISTINCT Name as LaunchProfileName, LOB as LineOfBusiness,CodeName,ExistingSKUProfile,Description,POMPOD as [POM/POD],CONVERT(varchar,LaunchDate,101) as LaunchDate,LaunchType,Regions as [Region(s)],CONVERT(varchar,AnnounceDate,101) as AnnounceDate,AnnounceFlag as [Announced(Y/N)],AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,CONVERT(varchar,PQSDate,101) as PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes FROM [launchmodeldev].[dbo].[FactLaunchProfiles] WHERE NAME = ?')
            params = (launchprofileparameter)
            if launchprofileparameter == "All" :
                id = cursor.execute(str(selectall))
            else :
                id = cursor.execute(str(selectone),params)
            for row in id.fetchall():
                row = list(row)
                print(row)
                my_sheet.append(row) 
    print("YES 200!")
    print(wb.sheetnames)
    return Response(save_virtual_workbook(wb),headers={'Content-Disposition': 'attatchment;','Content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})

@app.route('/launchplanning', methods=['GET', 'POST'])
def launchplans():
    launchplanparameter= request.args.get('launchplanparameter')
    print(launchplanparameter)

    if request.method=="GET":
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        results = []
        wb = Workbook()
        my_sheet = wb.active
        wb.title = "Launch-Profile-Template"
        
        #my_sheet = wb.create_sheet("Sheet_A")
        #my_sheet.title = "Launch-Profile-Template"
        #ws2 = wb.create_sheet("Sheet_B", 0)
        #ws2.title = "Title_B"

        my_sheet['A1'].value = "Launch-Profile-Template"
        my_sheet['A3'].value = "Required"
        my_sheet['B3'].value = "Required"
        my_sheet['C3'].value = "Required"
        my_sheet['D3'].value = "Required"
        my_sheet['F3'].value = "Required"
        my_sheet['G3'].value = "Required"
        my_sheet['H3'].value = "Required"
        my_sheet['I3'].value = "Required"
        my_sheet['K3'].value = "Required"
        my_sheet['A4'].value = "LaunchProfileName"
        my_sheet['B4'].value = "LineOfBusiness"
        my_sheet['C4'].value = "CodeName"
        my_sheet['D4'].value = "ExistingSKUProfile"
        my_sheet['E4'].value = "Description"
        my_sheet['F4'].value = "POM/POD"
        my_sheet['G4'].value = "LaunchDate"
        my_sheet['H4'].value = "LaunchType"
        my_sheet['I4'].value = "Region(s)"
        my_sheet['J4'].value = "AnnounceDate"
        my_sheet['K4'].value = "Announced(Y/N)"
        my_sheet['L4'].value = "AOCIPQ"
        my_sheet['M4'].value = "EOCIPQ"
        my_sheet['N4'].value = "APOCIPQ"
        my_sheet['O4'].value = "LOCIPQ"
        my_sheet['P4'].value = "FCCDate"
        my_sheet['Q4'].value = "DCVolume"
        my_sheet['R4'].value = "DTSVolume"
        my_sheet['S4'].value = "MSStoreIPQ"
        my_sheet['T4'].value = "Notes"
        my_sheet.title = "LaunchProfileTemplate"

        with conn.cursor() as cursor:
            selectall = text('SELECT DISTINCT Name as LaunchProfileName, LOB as LineOfBusiness,CodeName,ExistingSKUProfile,Description,POMPOD as [POM/POD],CONVERT(varchar,LaunchDate,101) as LaunchDate,LaunchType,Regions as [Region(s)],CONVERT(varchar,AnnounceDate,101) as AnnounceDate,AnnounceFlag as [Announced(Y/N)],AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,DCVolume,DTSVolume,MSStoreIPQ,Notes FROM [launchmodeldev].[dbo].[FactLaunchProfiles]')
            selectone = text('SELECT DISTINCT Name as LaunchProfileName, LOB as LineOfBusiness,CodeName,ExistingSKUProfile,Description,POMPOD as [POM/POD],CONVERT(varchar,LaunchDate,101) as LaunchDate,LaunchType,Regions as [Region(s)],CONVERT(varchar,AnnounceDate,101) as AnnounceDate,AnnounceFlag as [Announced(Y/N)],AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,DCVolume,DTSVolume,MSStoreIPQ,Notes FROM [launchmodeldev].[dbo].[FactLaunchProfiles] WHERE NAME = ?')
            id = cursor.execute(str(selectall))
            #if launchplanparameter == "All" :
            #    id = cursor.execute(str(selectall))
            #else :
            #    id = cursor.execute(str(selectone),params)
            for row in id.fetchall():
                row = list(row)
                print(row)
                my_sheet.append(row)
    print("YES 200!")
    print(wb.sheetnames)
    return Response(save_virtual_workbook(wb),headers={'Content-Disposition': 'attatchment;','Content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})

#filename=sheet.xlsx



    '''

     row = row + 1
            for i, row in enumerate(results[0]):
                my_sheet.cell(row=i, column=1).value = row

    #if not session.get("user"):
    #    return redirect(url_for("login"))
     if request.method == 'GET':
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config) 
        wb = Workbook
        with conn.cursor() as cursor: 
            select = text('SELECT DISTINCT Name as LaunchProfileName, LOB as LineOfBusiness,CodeName,ExistingSKUProfile,Description,POMPOD as [POM/POD],CONVERT(varchar,LaunchDate,101) as LaunchDate,LaunchType,Regions as [Region(s)],CONVERT(varchar,AnnounceDate,101) as AnnounceDate,AnnounceFlag as [Announced(Y/N)],AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,DCVolume,DTSVolume,MSStoreIPQ,Notes FROM [launchmodeldev].[dbo].[FactLaunchProfiles] WHERE NAME = ?')
            selectall = text('SELECT DISTINCT Origin,Destination,Customer,Channel,Other,DateType,TargetDate,Qty,FulfillmentScenario,NodeModeOne,NodeModeTwo,NodeModeThree,NodeModeFour,NodeModeFive,NodeModeSix FROM [launchmodeldev].[dbo].[PROD_LAUNCHPLANTEMPLATES]')
            #params = (launchprofileparameter)
            #if launchprofileparameter == "All" :
            id = cursor.execute(str(selectall))
            #else :
            #    id = cursor.execute(str(select),params)
            print(id)
            columns = [column[0] for column in id.description]
            print(columns)
            results = []
            for row in id.fetchall():
                print(row)
                results.append(dict(zip(columns, row)))
                #print(row)
        return jsonify(results) '''

@app.route('/launchparameters', methods=['GET', 'POST'])
def launchparameters():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method == 'GET':
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        with conn.cursor() as cursor:
            id = cursor.execute("SELECT DISTINCT * FROM [launchmodeldev].[dbo].[vw_LaunchPlans] Order by LProfileName ASC")
            columns = [column[0] for column in id.description]
            print(columns)
            results = []
            for row in id.fetchall():
                results.append(dict(zip(columns, row)))
            print(jsonify(results))
        return jsonify(results)

@app.route('/cascadingmeasures', methods=['GET', 'POST'])
def cascadingmeasures():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method == 'GET':
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        with conn.cursor() as cursor:
            id = cursor.execute("SELECT DISTINCT profiles.LProfileName,profiles.LProfileId,Launch.LaunchPlanName,Launch.LaunchPlanId,Launch.[Version] FROM [launchmodeldev].[dbo].[vw_LaunchPlans] as profiles LEFT JOIN [dbo].[FactLaunchPlans] as Launch  on  profiles.LProfileId = Launch.LaunchProfileId where Launch.LaunchPlanName is not null Order by profiles.LProfileName ASC")
            columns = [column[0] for column in id.description]
            print(columns)
            results = []
            for row in id.fetchall():
                results.append(dict(zip(columns, row)))
            print(jsonify(results))
        return jsonify(results)

@app.route('/launchskuattributes', methods=['GET', 'POST'])
def launchskuattributes():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method == 'GET':
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config,max=True)
        with conn.cursor() as cursor:
            id = cursor.execute(" SELECT DISTINCT PlanningBusinessUnit as Planningbusinessunit from dw.dimmaterial where IsDeviceActiveSku = 1 ")
            columns = [column[0] for column in id.description]
            print(columns)
            results = []
            for row in id.fetchall():
                results.append(dict(zip(columns, row)))
            print(jsonify(results))
        return jsonify(results)

@app.route('/launchversions', methods=['GET', 'POST'])
def launchversions():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method == 'GET':
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        with conn.cursor() as cursor:
            id = cursor.execute('SELECT distinct [version] from [dbo].[FactLaunchPlans]')
            columns = [column[0] for column in id.description]
            print(columns)
            results = []
            for row in id.fetchall():
                results.append(dict(zip(columns, row)))
            print(jsonify(results))
        return jsonify(results)

#Downloads
@app.route("/launchplandownloadfile", methods=['GET','POST'])
def launchplandownloadfile():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    launchplanparameter = request.args.get('launchplanparameter')
    print("API HIT")
    print(launchplanparameter)
    if request.method =='GET':
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        wb = Workbook()
        my_sheet = wb.active
        my_sheet2 = wb.create_sheet("Build Plan", 1)
        my_sheet.title = "Launch Plan"

        my_sheet['A1'].value = "Launch-Plan"
       
       
        my_sheet['F3'].value = "Required"
        my_sheet['G3'].value = "Required"
        my_sheet['H3'].value = "Required"
        
        my_sheet['A4'].value = "Origin"
        my_sheet['B4'].value = "Destination"
        my_sheet['C4'].value = "Customer"
        my_sheet['D4'].value = "Channel"
        my_sheet['E4'].value = "Other"
        my_sheet['F4'].value = "DateType"
        my_sheet['G4'].value = "TargetDate"
        my_sheet['H4'].value = "Qty"
        my_sheet['I4'].value = "FulfillmentScenario"
        my_sheet['J4'].value = "NodeModeOne"
        my_sheet['K4'].value = "NodeModeTwo"
        my_sheet['L4'].value = "NodeModeThree"
        my_sheet['M4'].value = "NodeModeFour"
        my_sheet['N4'].value = "NodeModeFive"
        my_sheet['O4'].value = "NodeModeSix"

        my_sheet2['A1'].value = "Build-Plan"
        my_sheet2['A3'].value = "Not Required"
        my_sheet2['B3'].value = "Not Required"
        my_sheet2['A4'].value = "Date"
        my_sheet2['B4'].value = "BuildQty"
        
        with conn.cursor() as cursor:
            selectlaunchplan = text('SELECT FLP.Origin, FLP.Destination,FLP.Customer,FLP.Channel,FLP.Other,FLP.DateType,CONVERT(varchar,FLP.TargetDate,101) as TargetDate,FLP.Qty, FLP.FulfillmentScenario,FLP.NodeModeOne,FLP.NodeModeTwo,FLP.NodeModeThree,FLP.NodeModeFour,FLP.NodeModeFive,FLP.NodeModeSix from [launchmodeldev].[dbo].[FactLaunchPlans] AS FLP WHERE EXISTS (SELECT TOP 1 [Version] FROM [launchmodeldev].[dbo].[FactLaunchPlans] as latest WHERE latest.LaunchPlanName = ? AND latest.Version = FLP.Version ORDER BY latest.ChangeDate DESC)')
            selectbuildplan = text('SELECT DISTINCT CAST(BPL.Date as date) as [Date],BPL.BuildQty FROM (SELECT TOP 1 [Version] from [launchmodeldev].[dbo].[FactLaunchPlans] WHERE LAUNCHPLANNAME = ? ORDER BY ChangeDate DESC ) AS FLP LEFT JOIN [dbo].[FactBuildPlans] AS BPL on FLP.[Version]=BPL.[Version]') 
            params = launchplanparameter
            launchplanid = cursor.execute(str(selectlaunchplan),params)
            for row in launchplanid.fetchall():
                row = list(row)
                print(row)
                my_sheet.append(row) 
            buildplanid = cursor.execute(str(selectbuildplan),params)
            for row in buildplanid.fetchall():
                row = list(row)
                print(row)
                my_sheet2.append(row)
    return Response(save_virtual_workbook(wb),headers={'Content-Disposition': 'attatchment;','Content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})


@app.route("/launchprofiledownloadfile", methods=['GET','POST'])
def launchprofiledownloadfile():
    #if not session.get("user"):
    #    return redirect(url_for("login"))

    launchprofileparameter = request.args.get('launchprofileparameter')
    
    if request.method == "GET":
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        results = []
        wb = Workbook()
        my_sheet = wb.active
        #wb.title = "Launch-Profile-Template"
        print(launchprofileparameter) #string "," of project names
        print("Workbook Created")
        paramlist = (launchprofileparameter.split(",")) #List of profile names
        length = int(len(paramlist))
        selectall = text('SELECT DISTINCT Name as LaunchProfileName, LOB as LineOfBusiness,CodeName,ExistingSKUProfile,Description,POMPOD as [POM/POD],CONVERT(varchar,LaunchDate,101) as LaunchDate,LaunchType,Regions as [Region(s)],CONVERT(varchar,AnnounceDate,101) as AnnounceDate,AnnounceFlag as [Announced(Y/N)],AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,CONVERT(varchar,PQSDate,101) as PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes FROM [launchmodeldev].[dbo].[FactLaunchProfiles]')
        selectone = text('SELECT DISTINCT Name as LaunchProfileName, LOB as LineOfBusiness,CodeName,ExistingSKUProfile,Description,POMPOD as [POM/POD],CONVERT(varchar,LaunchDate,101) as LaunchDate,LaunchType,Regions as [Region(s)],CONVERT(varchar,AnnounceDate,101) as AnnounceDate,AnnounceFlag as [Announced(Y/N)],AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,CONVERT(varchar,PQSDate,101) as PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes FROM [launchmodeldev].[dbo].[FactLaunchProfiles] WHERE NAME = ?')
        #i = 0
        my_sheet['A1'].value = "Launch-Profile-Template"
        my_sheet['A3'].value = "Required"
        my_sheet['B3'].value = "Required"
        my_sheet['C3'].value = "Required"
        my_sheet['D3'].value = "Required"
        my_sheet['F3'].value = "Required"
        my_sheet['G3'].value = "Required"
        my_sheet['H3'].value = "Required"
        my_sheet['I3'].value = "Required"
        my_sheet['K3'].value = "Required"
        my_sheet['A4'].value = "LaunchProfileName"
        my_sheet['B4'].value = "LineOfBusiness"
        my_sheet['C4'].value = "CodeName"
        my_sheet['D4'].value = "ExistingSKUProfile"
        my_sheet['E4'].value = "Description"
        my_sheet['F4'].value = "POM/POD"
        my_sheet['G4'].value = "LaunchDate"
        my_sheet['H4'].value = "LaunchType"
        my_sheet['I4'].value = "Region(s)"
        my_sheet['J4'].value = "AnnounceDate"
        my_sheet['K4'].value = "Announced(Y/N)"
        my_sheet['L4'].value = "AOCIPQ"
        my_sheet['M4'].value = "EOCIPQ"
        my_sheet['N4'].value = "APOCIPQ"
        my_sheet['O4'].value = "LOCIPQ"
        my_sheet['P4'].value = "FCCDate"
        my_sheet['Q4'].value = "PQSDate"
        my_sheet['R4'].value = "DCVolume"
        my_sheet['S4'].value = "DTSVolume"
        my_sheet['T4'].value = "MSStoreIPQ"
        my_sheet['U4'].value = "Notes"

        print(len(paramlist))
        print(length);
        print(type(length))

        with conn.cursor() as cursor:
            if 'All' in paramlist: #All Parameter give me everything
                id = cursor.execute(str(selectall))
                for row in id.fetchall():
                    row = list(row)
                    print(row)
                    my_sheet.append(row)
            else:
                for i in paramlist:
                    print(i)
                    print("we are looping!")
                    print(type(i))
                    print(i)
                    param = i
                    id = cursor.execute(str(selectone),param)
                    print("query executed!")
                    for row in id.fetchall():
                        row = list(row)
                        print(row)
                        my_sheet.append(row)
                        print("Success writing to file!")
                    
        cursor.close()
        
        if 'All' in paramlist:
            wb.title = "All-Launch Profiles"
            my_sheet.title = "All-Launch Profiles"
        elif len(paramlist) < 1:
            wb.title = str(paramlist[0]) + "- Launch Profile"
            my_sheet.title = str(paramlist[0]) +"- Launch Profile"
        elif len(paramlist) > 0:
            wb.title = "Multiple- Launch Profiles"
            my_sheet.title = "Multiple- Launch Profiles"

    print("YES 200!")
    print(wb.sheetnames)
    return Response(save_virtual_workbook(wb),headers={'Content-Disposition': 'attatchment;','Content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})



    if request.method=="POST":
        conn = getSQLConnection(app_config=app_config)
        f = request.files['fileupload']
        
        #print(f) 
        form = request.form
        FileName = f.filename
        launchID = form.get('launchprofilesDropdown')
        print(FileName, launchID)
        

        
        return redirect(url_for('launches'))


@app.route("/launchplandropdown", methods=['GET','POST'])
def launchplandropdown():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method == 'GET':
        data = request.get_json()
        conn = getSQLConnection(app_config=app_config)
        with conn.cursor() as cursor:
            id = cursor.execute("SELECT DISTINCT LAUNCHPLANNAME FROM [launchmodeldev].[dbo].[FactLaunchPlans]")
            columns = [column[0] for column in id.description]
            print(columns)
            results = []
            for row in id.fetchall():
                results.append(dict(zip(columns, row)))
            print(jsonify(results))
        return jsonify(results)


@app.route("/maxtest")

def maxtest():

    conn = getSQLConnection(app_config, max=True)

    with conn.cursor() as cursor:

            selectall = text('SELECT count(*) FROM dw.dimmaterial')

            id = cursor.execute(str(selectall))

            for row in id.fetchall():

                row = list(row)

                print(row)

    with conn.cursor() as cursor:

        selectall = text('SELECT count(*) FROM deliverods.shipmentdetails')

        id = cursor.execute(str(selectall))

        for row in id.fetchall():

            row = list(row)

            print(row)

    conn.close()

    return redirect(url_for('home'))



#Templates

@app.route('/launchplantemplate', methods=['GET', 'POST'])
def launchplantemplate():
    #if not session.get("user"):
    #return redirect(url_for("login"))
    if request.method == 'GET':
        conn = getSQLConnection(app_config=app_config)
        wb = Workbook()
        my_sheet =  wb.active
        my_sheet2 = wb.create_sheet("Build Plan - Template", 1)
        my_sheet.title = "Launch Plan - Template"
        my_sheet['A1'].value = "Launch-Plan-Template"

        my_sheet['F3'].value = "Required"
        my_sheet['G3'].value = "Required"
        my_sheet['H3'].value = "Required"

        my_sheet['A4'].value = "Origin"
        my_sheet['B4'].value = "Destination"
        my_sheet['C4'].value = "Customer"
        my_sheet['D4'].value = "Channel"
        my_sheet['E4'].value = "Other"
        my_sheet['F4'].value = "DateType"
        my_sheet['G4'].value = "TargetDate"
        my_sheet['H4'].value = "Qty"
        my_sheet['I4'].value = "FulfillmentScenario"
        my_sheet['J4'].value = "NodeModeOne"
        my_sheet['K4'].value = "NodeModeTwo"
        my_sheet['L4'].value = "NodeModeThree"
        my_sheet['M4'].value = "NodeModeFour"
        my_sheet['N4'].value = "NodeModeFive"
        my_sheet['O4'].value = "NodeModeSix"

        my_sheet2['A1'].value = "Build-Plan-Template"
        my_sheet2['A3'].value = "Not Required"
        my_sheet2['B3'].value = "Not Required"
        my_sheet2['A4'].value = "Date"
        my_sheet2['B4'].value = "BuildQty"

        with conn.cursor() as cursor:
            selectlaunchplan = text('SELECT DISTINCT [Origin - Port] as Origin,Destination,Customer,Channel,Other,DateType,CONVERT(varchar,TargetDate,101) as TargetDate,Qty, FulfillmentScenario,NodeModeOne,NodeModeTwo,NodeModeThree,NodeModeFour,NodeModeFive,NodeModeSix from dbo.FactLaunchMasterPlanTemplate')
            
            launchplanid = cursor.execute(str(selectlaunchplan))
            for row in launchplanid.fetchall():
                row = list(row)
                print(row)
                my_sheet.append(row) 
     
    return Response(save_virtual_workbook(wb),headers={'Content-Disposition': 'attatchment;','Content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})
  
@app.route('/launchprofiletemplate', methods=['GET', 'POST'])
def launchprofiletemplate():
    #if not session.get("user"):
    #return redirect(url_for("login"))
    if request.method == 'GET':
        wb = Workbook()
        my_sheet = wb.active
        wb.title = "Launch-Profile-Template"
        my_sheet['A1'].value = "Launch-Profile-Template"
        my_sheet['A3'].value = "Required"
        my_sheet['B3'].value = "Required"
        my_sheet['C3'].value = "Required"
        my_sheet['D3'].value = "Required"
        my_sheet['F3'].value = "Required"
        my_sheet['G3'].value = "Required"
        my_sheet['H3'].value = "Required"
        my_sheet['I3'].value = "Required"
        my_sheet['K3'].value = "Required"
        my_sheet['A4'].value = "LaunchProfileName"
        my_sheet['B4'].value = "LineOfBusiness"
        my_sheet['C4'].value = "CodeName"
        my_sheet['D4'].value = "ExistingSKUProfile"
        my_sheet['E4'].value = "Description"
        my_sheet['F4'].value = "POM/POD"
        my_sheet['G4'].value = "LaunchDate"
        my_sheet['H4'].value = "LaunchType"
        my_sheet['I4'].value = "Region(s)"
        my_sheet['J4'].value = "AnnounceDate"
        my_sheet['K4'].value = "Announced(Y/N)"
        my_sheet['L4'].value = "AOCIPQ"
        my_sheet['M4'].value = "EOCIPQ"
        my_sheet['N4'].value = "APOCIPQ"
        my_sheet['O4'].value = "LOCIPQ"
        my_sheet['P4'].value = "FCCDate"
        my_sheet['Q4'].value = "DTSVolume"
        my_sheet['R4'].value = "DCVolume"
        my_sheet['S4'].value = "MSStoreIPQ"
        my_sheet['T4'].value = "Notes"
        my_sheet.title = "LaunchProfileTemplate"
        row = ['SampleLaunchProfile', 'Surface Devices', 'ProjectX', 'Surface Studio', 'Description', 'POM', '05/14/2021', 'MSD', 'AOC;EOC;APOC;', '05/20/2021', 'Y', 12, 12, 12, 12, '05/20/2021', 12, 12, 12, 'Notes']
        my_sheet.append(row)
        print("YES 200!")
        print(wb.sheetnames)
    return Response(save_virtual_workbook(wb),headers={'Content-Disposition': 'attatchment;','Content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})



#Uploads
#Uploads for excel and Launch Date Doing weird String bullshit

@app.route("/uploadlaunchprofilefile", methods=['GET','POST'])
def uploadlaunchprofilefile():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method=="POST":
        conn = getSQLConnection(app_config=app_config)
        f = request.files['fileupload']
        rows = []
        wb2 = load_workbook(f)
        sh = wb2.active
        df = pd.DataFrame(wb2.active.values)
        launchprofiledf = df.iloc[4:]
        launchprofiledf.columns = ['Name',"LOB",
        "CodeName","ExistingSKUProfile",
        "Description","POMPOD",
        "LaunchDate","LaunchType",
        "Regions","AnnounceDate",
        "AnnounceFlag","AOCIPQ",
        "EOCIPQ","APOCIPQ",
        "LOCIPQ","FCCDate","PQSDate","DCVolume",
        "DTSVolume","MSStoreIPQ",
        "Notes"]
        launchprofiledf["ChangeDate"] = datetime.now()
        launchprofiledf["CreatedBy"] = 'Chosbo@microsoft.com'
        #launchprofiledf['LaunchDate'] = pd.to_datetime(launchprofiledf['LaunchDate'])
        #launchprofiledf['ChangeDate'] = pd.to_datetime(launchprofiledf['ChangeDate'])
        #launchprofiledf['AnnounceDate'] = pd.to_datetime(launchprofiledf['AnnounceDate'])
        #launchprofiledf['FCCDate'] = pd.to_datetime(launchprofiledf['FCCDate'])
        #launchprofiledf = launchprofiledf.fillna(value='N/A')


        print("Lets look at this")
        #print(launchprofiledf['AnnounceDate'])
        #print(launchprofiledf['LaunchDate'])
        #print(launchprofiledf['ChangeDate'])
        #print(launchprofiledf['FCCDate'])
        #print(launchprofiledf)
        with conn.cursor() as cursor: 
            id = cursor.execute("SELECT DISTINCT Id, Name, LOB,CodeName,ExistingSKUProfile,Description,POMPOD,convert(varchar,LaunchDate,22) as LaunchDate,LaunchType,Regions,convert(varchar,AnnounceDate,22) as AnnounceDate,AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,CONVERT(varchar,PQSDate,101) as PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes,convert(varchar,ChangeDate,22) as ChangeDate, CreatedBy FROM [launchmodeldev].[dbo].[FactLaunchProfiles]")
            result = id.fetchall()
            #pd.option_context('display.max_rows',None,'display.max_columns',None,'display.precision',3)
            df2 = list(result)
            df2 = pd.DataFrame.from_records(df2, columns = ['Id','Name','LOB','CodeName','ExistingSKUProfile','Description','POMPOD','LaunchDate','LaunchType','Regions','AnnounceDate','AnnounceFlag','AOCIPQ','EOCIPQ','APOCIPQ','LOCIPQ','FCCDate','PQSDate','DCVolume','DTSVolume','MSStoreIPQ','Notes','ChangeDate','CreatedBy'])
            #display(df2)
            df2 = pd.DataFrame.from_records(df2, columns = ['Id','Name']) #Current Profiles Entered in Tool
            df = pd.merge(launchprofiledf,df2,on='Name',how='left') #Uploaded Launches that have matches from database
            df['Id'] = df['Id'].fillna(value='N/A') #All non-matches get assigned a 'N/A' value for ID
            #df = df.fillna(value='N/A')
            print("Merged Dataset")
            print(df.keys())
            newdf = df.loc[(df.Id == 'N/A')] #All New Profiles
            df = df.loc[(df.Id != 'N/A')] #All Existing Profiles in database

          
            newdf['LaunchDate'] = pd.to_datetime(newdf['LaunchDate'])
            newdf['FCCDate'] = pd.to_datetime(newdf['FCCDate'])
            newdf['PQSDate'] = pd.to_datetime(newdf['PQSDate'])
            newdf['ChangeDate'] = pd.to_datetime(newdf['ChangeDate'])
            newdf['AnnounceDate'] = pd.to_datetime(newdf['AnnounceDate'])


            newdf['FCCDate'] = (newdf['FCCDate'].astype(str).replace({'NaT': None}))
            newdf['PQSDate'] = (newdf['PQSDate'].astype(str).replace({'NaT': None}))
            newdf['AnnounceDate'] = (newdf['AnnounceDate'].astype(str).replace({'NaT': None}))              
            newdf['LaunchDate'] = (newdf['LaunchDate'].astype(str).replace({'NaT': None})) 

            df['LaunchDate'] = pd.to_datetime(df['LaunchDate'])
            df['FCCDate'] = pd.to_datetime(df['FCCDate'])
            df['PQSDate'] = pd.to_datetime(df['PQSDate'])
            df['ChangeDate'] = pd.to_datetime(df['ChangeDate'])
            df['AnnounceDate'] = pd.to_datetime(df['AnnounceDate'])

            df['FCCDate'] = (df['FCCDate'].astype(str).replace({'NaT': None}))
            df['PQSDate'] = (df['PQSDate'].astype(str).replace({'NaT': None}))
            df['AnnounceDate'] = (df['AnnounceDate'].astype(str).replace({'NaT': None}))              
            df['LaunchDate'] = (df['LaunchDate'].astype(str).replace({'NaT': None})) 
           

  
            #updatetext = text("UPDATE [launchmodeldev].[dbo].[FactLaunchProfiles] SET Name=?,LOB=?,CodeName=?,ExistingSKUProfile=?,Description=?,POMPOD=?,LaunchDate=?,LaunchType=?,Regions=?,AnnounceDate=?,AnnaounceFlag=?,AOCIPQ=?,EOCIPQ=?,APOCIPQ=?,LOCIPQ=?,FCCDate=?,DCVolume=?,DTSVolume=?,MSStoreIPQ=?,Notes=?,ChangeDate=?,CreatedBy=? where Name=?")
            deletetext = text("DELETE FROM [launchmodeldev].[dbo].[FactLaunchProfiles] WHERE Id = ?")
            inserttext = text("INSERT INTO [launchmodeldev].[dbo].[FactLaunchProfiles](Id,Name,LOB,CodeName,ExistingSKUProfile,Description, POMPOD,LaunchDate,LaunchType,Regions,AnnounceDate,AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,FCCDate,PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes,ChangeDate,CreatedBy) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)")
            newrecords = text("INSERT INTO [launchmodeldev].[dbo].[FactLaunchProfiles](Id,Name,LOB,CodeName,ExistingSKUProfile,Description, POMPOD,LaunchDate,LaunchType,Regions,AnnounceDate,AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,FCCDate,PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes,ChangeDate,CreatedBy) VALUES(NEWID(),?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)")
            
            if df.empty == False:
                print("we hit df")
                for row in df.itertuples():
                    params = (row.Id)
                    cursor.execute(str(deletetext),params)
                    params = (row.Id,row.Name,row.LOB,row.CodeName,row.ExistingSKUProfile,row.Description,row.POMPOD,row.LaunchDate,row.LaunchType,row.Regions,row.AnnounceDate,row.AnnounceFlag,row.AOCIPQ,row.EOCIPQ,row.APOCIPQ,row.LOCIPQ,row.FCCDate,row.PQSDate,row.DCVolume,row.DTSVolume,row.MSStoreIPQ,row.Notes,row.ChangeDate,row.CreatedBy)
                    cursor.execute(str(inserttext),params)
            if newdf.empty == False:
                print("we hit newdf")
                for row in newdf.itertuples():
                    params = (row.Name,row.LOB,row.CodeName,row.ExistingSKUProfile,row.Description,row.POMPOD,row.LaunchDate,row.LaunchType,row.Regions,row.AnnounceDate,row.AnnounceFlag,row.AOCIPQ,row.EOCIPQ,row.APOCIPQ,row.LOCIPQ,row.FCCDate,row.PQSDate,row.DCVolume,row.DTSVolume,row.MSStoreIPQ,row.Notes,row.ChangeDate,row.CreatedBy)
                    cursor.execute(str(newrecords),params)
        
        cursor.close()
        
        return redirect(url_for('launches'))


@app.route("/uploadlaunchplanfile", methods=['GET','POST'])
def uploadlaunchplanfile():
    #if not session.get("user"):
    #    return redirect(url_for("login"))

    if request.method=="POST":
        conn = getSQLConnection(app_config=app_config)
        f = request.files['fileupload']
        rows = []
        wb2 = load_workbook(f)
        form = request.form
        FileName = f.filename
        ChangeDate = datetime.now()
        versionparameter = str(FileName) + " - " + str(ChangeDate)
        launchID = form.get('launchprofilesDropdown')
        print(FileName, launchID)
        print(wb2.sheetnames)
        launchsheet = wb2.sheetnames[0]
        buildsheet = wb2.sheetnames[1]
        launchplandf = pd.DataFrame(wb2[launchsheet].values)
        launchplandf = launchplandf.iloc[4:]
        launchplandf.columns = ['Origin','Destination',"Customer","Channel","Other","DateType","TargetDate","Qty","FulfillmentScenario","NodeModeOne","NodeModeTwo","NodeModeThree","NodeModeFour","NodeModeFive","NodeModeSix"]

        #print(launchplandf)

        buildplandf = pd.DataFrame(wb2[buildsheet].values)
        buildplandf = buildplandf.iloc[4:]
        buildplandf.columns = ['Date','BuildQty']


        #print(buildplandf)

      
        id = uuid.uuid1()
        #print("Add Columns to DF")
        launchplandf["LaunchPlanId"] = str(id)
        launchplandf["ChangeDate"] = ChangeDate
        launchplandf["UpdatedBy"] = 'Chosbo@microsoft.com'
        launchplandf["LaunchProfileId"] = launchID
        launchplandf["LaunchPlanName"] = FileName
        launchplandf["Version"] = str(FileName) + " - " + str(ChangeDate)
        launchplandf = launchplandf.fillna(value='N/A')
        #launchplandf['TargetDate'] = launchplandf.to_datetime(launchplandf['TargetDate'],format="%m/%d/%Y, %H:%M:%S",errors='raise')
        #launchplandf['ChangeDate'] = launchplandf.to_datetime(launchplandf['ChangeDate'],format="%m/%d/%Y, %H:%M:%S",errors='raise')
        launchplandf['TargetDate'] = pd.to_datetime(launchplandf['TargetDate'])
        launchplandf['ChangeDate'] = pd.to_datetime(launchplandf['ChangeDate'])



        buildplandf["LaunchPlanId"] = str(id)
        buildplandf["ChangeDate"] = ChangeDate
        buildplandf["UpdatedBy"] = 'Chosbo@microsoft.com'
        buildplandf["LaunchPlanName"] = FileName
        buildplandf["Version"] = str(FileName) + " - " + str(ChangeDate)
        buildplandf = buildplandf.fillna(value='N/A')
        buildplandf['ChangeDate'] = pd.to_datetime(buildplandf['ChangeDate'])
        #buildplandf['ChangeDate'] = buildplandf.to_datetime(buildplandf['ChangeDate'],format="%m/%d/%Y, %H:%M:%S",errors='raise')


        
        print(launchplandf)
        print(buildplandf)
        with conn.cursor() as cursor:
            #id = cursor.execute("SELECT DISTINCT Name, LOB,CodeName,ExistingSKUProfile,Description,POMPOD,convert(varchar,LaunchDate,22) as LaunchDate,LaunchType,Regions,convert(varchar,AnnounceDate,22) as AnnounceDate,AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,DCVolume,DTSVolume,MSStoreIPQ,Notes,convert(varchar,ChangeDate,22) as ChangeDate, CreatedBy FROM [launchmodeldev].[dbo].[FactLaunchProfiles]")
            #result = id.fetchall()
            #pd.option_context('display.max_rows',None,'display.max_columns',None,'display.precision',3)
            #df2 = list(result)
            #df2 = pd.DataFrame.from_records(df2, columns = ['Name','LOB','CodeName','ExistingSKUProfile','Description','POMPOD','LaunchDate','LaunchType','Regions','AnnounceDate','AnnounceFlag','AOCIPQ','EOCIPQ','APOCIPQ','LOCIPQ','DCVolume','DTSVolume','MSStoreIPQ','Notes','ChangeDate','CreatedBy'])
            #display(df2)
            #deletetext = text("DELETE FROM [launchmodeldev].[dbo].[FactLaunchProfiles] WHERE NAME = ?")
            inserttext = text("INSERT INTO [launchmodeldev].[dbo].[FactLaunchPlans](LaunchPlanId,LaunchProfileId,LaunchPlanName,ChangeDate,Version,UpdatedBy,Origin,Destination,Customer,Channel,Other,DateType,TargetDate,Qty,FulfillmentScenario,NodeModeOne,NodeModeTwo,NodeModeThree,NodeModeFour,NodeModeFive,NodeModeSix)VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)")
            insertbuildplan = text("INSERT INTO [launchmodeldev].[dbo].[FactBuildPlans](LaunchPlanId,ChangeDate,Version,UpdatedBy,Date,BuildQty)VALUES(?,?,?,?,?,?)")
            procedure = text("EXEC sp_GenerateShipPlan @Version = ?")
            
            #paramspro = 'poop2'
            #cursor.execute(str(procedure),paramspro)
            #df3 = pd.merge(df2,df,on='Name')
            #display(df3)
            #df3.to_csv(r'C:\Users\chosbo\Desktop\testdf.csv',index=False,header=True)
            for row in launchplandf.itertuples():
                print(row)
                params =(row.LaunchPlanId,row.LaunchProfileId,row.LaunchPlanName,row.ChangeDate,row.Version,row.UpdatedBy,row.Origin,row.Destination,row.Customer,row.Channel,row.Other,row.DateType,row.TargetDate,row.Qty,row.FulfillmentScenario,row.NodeModeOne,row.NodeModeTwo,row.NodeModeThree,row.NodeModeFour,row.NodeModeFive,row.NodeModeSix)
                cursor.execute(str(inserttext),params)
            for row in buildplandf.itertuples():
                #print(row)
                params =(row.LaunchPlanId,row.ChangeDate,row.Version,row.UpdatedBy,row.Date,row.BuildQty)
                cursor.execute(str(insertbuildplan),params)
            cursor.execute(str(procedure),versionparameter)
        cursor.close()


            #result = ([dict(zip(i.keys(),i.values())) for i in id])
            #df2 = pd.DataFrame(result)
            #print(df2)
            #columns = [column[0] for column in id.description]
            #print(columns)
            #results = []
            #for row in id.fetchall():
            #    results.append(dict(zip(columns, row)))
            #    jsonobject = jsonify(results)
            #print(jsonobject[0])
        #insert = text("MERGE [launchmodeldev].[dbo].[FactLaunchProfiles] as mrg USING (SELECT :Id,:Name,:LOB,:CodeName,:ExistingSKUProfile,:Description,:POMPOD,:LaunchDate,:LaunchType,:Regions,:AnnounceDate,:AnnouceFlag,:AOCIPQ,:EOCIPQ,:APOCIPQ,:LOCIPQ,:DCVolume,:DTSVolume,:MSStoreIPQ,:Notes,:ChangeDate,:CreatedBy) as upload  ON mrg.Name = upload.Name WHEN MATCHED UPDATE SET msg.Name=upload.Name,msg.LOB=upload.LOB,msg.CodeName=upload.CodeName,msg.ExistingSKUProfile=upload.ExistingSKUProfile,msg.Description=upload.Description,msg.POMPOD=upload.POMPOD,msg.LaunchDate=upload.LaunchDate,msg.LaunchType=upload.LaunchType,msg.Regions=upload.Regions,msg.AnnounceDate=upload.AnnounceDate,msg.AnnouceFlag=upload.AnnouceFlag,msg.AOCIPQ=upload.AOCIPQ,msg.EOCIPQ=upload.EOCIPQ,msg.APOCIPQ=upload.APOCIPQ,msg.LOCIPQ=upload.LOCIPQ,msg.DCVolume=upload.DCVolume,msg.DTSVolume=upload.DTSVolume,msg.MSStoreIPQ=upload.MSStoreIPQ,msg.Notes=upload.Notes, msg.ChangeDate=upload.ChangeDate, msg.Createdby=upload.Createdby WHEN NOT MATCHED THEN INSERT(Id,Name,LOB,CodeName,ExistingSKUProfile,Description,POMPOD,LaunchDate,LaunchType,Regions,AnnounceDate,AnnouceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,DCVolume,DTSVolume,MSStoreIPQ,Notes,ChangeDate,CreatedBy)VALUES(upload.ID,upload.Name,upload.LOB,upload.CodeName,upload.ExistingSKUProfile,upload.Description,upload.POMPOD,upload.LaunchDate,upload.LaunchType,upload.Regions,upload.AnnounceDate,upload.AnnouceFlag,upload.AOCIPQ,upload.EOCIPQ,upload.APOCIPQ,upload.LOCIPQ,upload.DCVolume,upload.DTSVolume,upload.MSStoreIPQ,upload.Notes,upload.ChangeDate,upload.CreatedBy)")
        #with conn.cursor() as cursor:
        #    for row in df.itertuples(): #iterrows():#itertuples()
        #        print(row)
        #        #params = (FileName,id,launchID,row.LaunchPlanName,str(datetime.now()),row.Version,row.UpdatedBy,row.Origin_Location,row.Destination_Country,row.Customer,row.Channel,row.Other,row.Date_Type,str(row.Target_Date),row.Qty,row.Fulfillment_Scenario,row.Model_Type,row.NodeModeOne,row.NodeModeTwo,row.NodeModeThree,row.NodeModeFour,row.NodeModeFive,row.NodeModeSix)
        #        cursor.execute(insert,['sadfsdf'=Id,row.Name:Name,row.CodeName:CodeName,row.ExistingSKUProfile:ExistingSKUProfile,row.Description:Description,row.POMPOD:POMPOD,row.LaunchDate:LaunchDate,row.LaunchType:LaunchType,row.Regions:Regions,row.AnnounceDate:AnnounceDate,row.AnnouceFlag:AnnouceFlag,row.AOCIPQ:AOCIPQ,row.EOCIPQ:EOCIPQ,row.APOCIPQ:APOCIPQ,row.LOCIPQ:LOCIPQ,row.DCVolume:DCVolume,row.DTSVolume:DTSVolume,row.MSStoreIPQ:MSStoreIPQ,row.Notes:Notes,datetime.now():ChangeDate,"Chosbo":CreatedBy])
        #cursor.close()
        #print(df)
    return render_template('launchplans.html')


# HOT-FIX: Have to figure out how to add / preserve the ID of the table object being deleted and updated.
@app.route("/uploadfile", methods=['GET','POST'])
def uploadFiles():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method=="POST":
        conn = getSQLConnection(app_config=app_config)
        f = request.files['fileupload']
        form = request.form
        FileName = f.filename
        launchID = form.get('launchprofilesDropdown')
        id = uuid.uuid1()
        print(FileName, launchID)
        for key in form.keys():
            print(FileName,launchID)
            for value in form.getlist(key):
                print(key,":",value)
        fstring = f.read()
        id = uuid.uuid1()
        text_obj = fstring.decode('UTF-8')
        data = io.StringIO(text_obj)
        df = pd.read_csv(data,sep=",")
        df = df.iloc[2:]
        #print(df) 
        df.columns=['Name',"LOB",
        "CodeName","ExistingSKUProfile",
        "Description","POMPOD",
        "LaunchDate","LaunchType",
        "Regions","AnnounceDate",
        "AnnounceFlag","AOCIPQ",
        "EOCIPQ","APOCIPQ",
        "LOCIPQ","FCCDate","PQSDate","DCVolume",
        "DTSVolume","MSStoreIPQ",
        "Notes"] 
        df["ChangeDate"] = datetime.now()
        df["CreatedBy"] = 'Chosbo@microsoft.com'
        df = df.fillna(value='N/A')
        #print(df)

        with conn.cursor() as cursor: 
            id = cursor.execute("SELECT DISTINCT Id, Name, LOB,CodeName,ExistingSKUProfile,Description,POMPOD,convert(varchar,LaunchDate,22) as LaunchDate,LaunchType,Regions,convert(varchar,AnnounceDate,22) as AnnounceDate,AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,CONVERT(varchar,FCCDate,101) as FCCDate,CONVERT(varchar,PQSDate,101) as PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes,convert(varchar,ChangeDate,22) as ChangeDate, CreatedBy FROM [launchmodeldev].[dbo].[FactLaunchProfiles]")
            result = id.fetchall()
            #pd.option_context('display.max_rows',None,'display.max_columns',None,'display.precision',3)
            df2 = list(result)
            df2 = pd.DataFrame.from_records(df2, columns = ['Id','Name','LOB','CodeName','ExistingSKUProfile','Description','POMPOD','LaunchDate','LaunchType','Regions','AnnounceDate','AnnounceFlag','AOCIPQ','EOCIPQ','APOCIPQ','LOCIPQ','FCCDate','PQSDate','DCVolume','DTSVolume','MSStoreIPQ','Notes','ChangeDate','CreatedBy'])
            #display(df2)
            df2 = pd.DataFrame.from_records(df2, columns = ['Id','Name'])
            df = pd.merge(df,df2,on='Name',how='left')
            df = df.fillna(value='N/A')
            newdf = df.loc[(df.Id == 'N/A')]
            df = df.loc[(df.Id != 'N/A')] #& (df.carrier == "B6")]
            print(df.columns)
            print(df)
            print(newdf)
            updatetext = text("UPDATE [launchmodeldev].[dbo].[FactLaunchProfiles] SET Name=?,LOB=?,CodeName=?,ExistingSKUProfile=?,Description=?,POMPOD=?,LaunchDate=?,LaunchType=?,Regions=?,AnnounceDate=?,AnnaounceFlag=?,AOCIPQ=?,EOCIPQ=?,APOCIPQ=?,LOCIPQ=?,FCCDate=?,DCVolume=?,DTSVolume=?,MSStoreIPQ=?,Notes=?,ChangeDate=?,CreatedBy=? where Name=?")
            deletetext = text("DELETE FROM [launchmodeldev].[dbo].[FactLaunchProfiles] WHERE Id = ?")
            inserttext = text("INSERT INTO [launchmodeldev].[dbo].[FactLaunchProfiles](Id,Name,LOB,CodeName,ExistingSKUProfile,Description, POMPOD,LaunchDate,LaunchType,Regions,AnnounceDate,AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,FCCDate,PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes,ChangeDate,CreatedBy) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)")
            newrecords = text("INSERT INTO [launchmodeldev].[dbo].[FactLaunchProfiles](Id,Name,LOB,CodeName,ExistingSKUProfile,Description, POMPOD,LaunchDate,LaunchType,Regions,AnnounceDate,AnnounceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,FCCDate,PQSDate,DCVolume,DTSVolume,MSStoreIPQ,Notes,ChangeDate,CreatedBy) VALUES(NEWID(),?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)")
            
            if df.empty == False:
                for row in df.itertuples():
                    print(row.Name)
                    params = (row.Id)
                    cursor.execute(str(deletetext),params)
                    params = (row.Id,row.Name,row.LOB,row.CodeName,row.ExistingSKUProfile,row.Description,row.POMPOD,row.LaunchDate,row.LaunchType,row.Regions,row.AnnounceDate,row.AnnounceFlag,row.AOCIPQ,row.EOCIPQ,row.APOCIPQ,row.LOCIPQ,row.FCCDate,row.PQSDate,row.DCVolume,row.DTSVolume,row.MSStoreIPQ,row.Notes,row.ChangeDate,row.CreatedBy)
                    cursor.execute(str(inserttext),params)
            elif newdf.empty == False:
                for row in newdf.itertuples():
                    params = (row.Name,row.LOB,row.CodeName,row.ExistingSKUProfile,row.Description,row.POMPOD,row.LaunchDate,row.LaunchType,row.Regions,row.AnnounceDate,row.AnnounceFlag,row.AOCIPQ,row.EOCIPQ,row.APOCIPQ,row.LOCIPQ,row.FCCDate,row.PQSDate,row.DCVolume,row.DTSVolume,row.MSStoreIPQ,row.Notes,row.ChangeDate,row.CreatedBy)
                    cursor.execute(str(newrecords),params)
            else:
                cursor.close()
        return redirect(url_for('launches'))

            #result = ([dict(zip(i.keys(),i.values())) for i in id])
            #df2 = pd.DataFrame(result)
            #print(df2)
            #columns = [column[0] for column in id.description]
            #print(columns)
            #results = []
            #for row in id.fetchall():
            #    results.append(dict(zip(columns, row)))
            #    jsonobject = jsonify(results)
            #print(jsonobject[0])
        #insert = text("MERGE [launchmodeldev].[dbo].[FactLaunchProfiles] as mrg USING (SELECT :Id,:Name,:LOB,:CodeName,:ExistingSKUProfile,:Description,:POMPOD,:LaunchDate,:LaunchType,:Regions,:AnnounceDate,:AnnouceFlag,:AOCIPQ,:EOCIPQ,:APOCIPQ,:LOCIPQ,:DCVolume,:DTSVolume,:MSStoreIPQ,:Notes,:ChangeDate,:CreatedBy) as upload  ON mrg.Name = upload.Name WHEN MATCHED UPDATE SET msg.Name=upload.Name,msg.LOB=upload.LOB,msg.CodeName=upload.CodeName,msg.ExistingSKUProfile=upload.ExistingSKUProfile,msg.Description=upload.Description,msg.POMPOD=upload.POMPOD,msg.LaunchDate=upload.LaunchDate,msg.LaunchType=upload.LaunchType,msg.Regions=upload.Regions,msg.AnnounceDate=upload.AnnounceDate,msg.AnnouceFlag=upload.AnnouceFlag,msg.AOCIPQ=upload.AOCIPQ,msg.EOCIPQ=upload.EOCIPQ,msg.APOCIPQ=upload.APOCIPQ,msg.LOCIPQ=upload.LOCIPQ,msg.DCVolume=upload.DCVolume,msg.DTSVolume=upload.DTSVolume,msg.MSStoreIPQ=upload.MSStoreIPQ,msg.Notes=upload.Notes, msg.ChangeDate=upload.ChangeDate, msg.Createdby=upload.Createdby WHEN NOT MATCHED THEN INSERT(Id,Name,LOB,CodeName,ExistingSKUProfile,Description,POMPOD,LaunchDate,LaunchType,Regions,AnnounceDate,AnnouceFlag,AOCIPQ,EOCIPQ,APOCIPQ,LOCIPQ,DCVolume,DTSVolume,MSStoreIPQ,Notes,ChangeDate,CreatedBy)VALUES(upload.ID,upload.Name,upload.LOB,upload.CodeName,upload.ExistingSKUProfile,upload.Description,upload.POMPOD,upload.LaunchDate,upload.LaunchType,upload.Regions,upload.AnnounceDate,upload.AnnouceFlag,upload.AOCIPQ,upload.EOCIPQ,upload.APOCIPQ,upload.LOCIPQ,upload.DCVolume,upload.DTSVolume,upload.MSStoreIPQ,upload.Notes,upload.ChangeDate,upload.CreatedBy)")
        #with conn.cursor() as cursor:
        #    for row in df.itertuples(): #iterrows():#itertuples()
        #        print(row)
        #        #params = (FileName,id,launchID,row.LaunchPlanName,str(datetime.now()),row.Version,row.UpdatedBy,row.Origin_Location,row.Destination_Country,row.Customer,row.Channel,row.Other,row.Date_Type,str(row.Target_Date),row.Qty,row.Fulfillment_Scenario,row.Model_Type,row.NodeModeOne,row.NodeModeTwo,row.NodeModeThree,row.NodeModeFour,row.NodeModeFive,row.NodeModeSix)
        #        cursor.execute(insert,['sadfsdf'=Id,row.Name:Name,row.CodeName:CodeName,row.ExistingSKUProfile:ExistingSKUProfile,row.Description:Description,row.POMPOD:POMPOD,row.LaunchDate:LaunchDate,row.LaunchType:LaunchType,row.Regions:Regions,row.AnnounceDate:AnnounceDate,row.AnnouceFlag:AnnouceFlag,row.AOCIPQ:AOCIPQ,row.EOCIPQ:EOCIPQ,row.APOCIPQ:APOCIPQ,row.LOCIPQ:LOCIPQ,row.DCVolume:DCVolume,row.DTSVolume:DTSVolume,row.MSStoreIPQ:MSStoreIPQ,row.Notes:Notes,datetime.now():ChangeDate,"Chosbo":CreatedBy])
        #cursor.close()
        #print(df)
    

'''

      
@app.route('/uploadfile', methods=['POST'])
def uploadFiles():
    if request.method=="POST":
        f = request.files['fileupload']
        test = request.form['form-select']
        print(test)
        form = request.form
        for key in form.keys():
            for value in form.getlist(key):
                print(key,":",value)
        fstring = f.read()
        text_obj = fstring.decode('UTF-8')
        data = io.StringIO(text_obj)
        df = pd.read_csv(data,sep=",")
        #csv_dicts = [{k: v for k, v in row.items()} for row in csv.DictReader(fstring.splitlines(), skipinitialspace=True)]  
        print(df)
    return("Success") 
    @app.route('/launchprofiles',methods=['GET','POST'])
    def launchprofiles():
        if request.method == "GET":
            mylist = []
            conn = getSQLConnection(app_config=app_config)
            with conn.cursor() as cursor:
                id = cursor.execute("SELECT DISTINCT ID,Name FROM [launchmodeldev].[dbo].[FactLaunchProfiles]")
                result = id.fetchall()
            print(result)
            return str(result)
        #for launch in launches:
        #   launch = {'LaunchName':launch['name']}
        #   mylist.append(launch)
        # 


launches = [
    {
        'launchName':"Spring Launch",
        "sku":[
            {
            "skuName": "Surface Laptop",
            "price": 109
            }
        ]
    }
]

launches = [
    {
        'LaunchProfile':"id",
        "data":[
            {
            "name": "FalconX",
            "launchdate": "12/12/12",
            "changedate": "12/12/12"
            }

        ]
    }
]



#Post, add a launch
@app.route('/launchprofiles',methods=['POST'])
def create_launch():
    request_data = request.get_json()
    #new_launch = {'launchName':request_data['data']}
    return request_data

launches = [
    
    {'id': 12,
    'name':'FalconX'}, 
    {'id': 13,
    'name':'FalconB'}
]

@app.route('/launch/<string:name>/sku')
def get_sku_in_launch(name):
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    for launch in launches:
        if launch["launchName"] == name:
            return jsonify({'sku':launch['sku']})
    return jsonify({"message":"Sku Not Found"})
@app.route('/launchprofiles',methods=['GET'])
def get_allLaunches():
    mylist = []
    for launch in launches:
        launch = {'LaunchName':launch['name']}
        mylist.append(launch)

    return jsonify(mylist)




@app.route("/uploadfile", methods=['POST'])
def uploadFiles():
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    if request.method=="POST":
        conn = getSQLConnection(app_config=app_config)
        f = request.files['fileupload']
        print(f) 
        form = request.form
        FileName = f.filename
        launchID = form.get('launchprofilesDropdown')
        print(FileName, launchID)
        for key in form.keys():
            print(FileName,launchID)
            for value in form.getlist(key):
                print(key,":",value)
        fstring = f.read()
        id = uuid.uuid1()
        text_obj = fstring.decode('UTF-8')
        data = io.StringIO(text_obj)
        df = pd.read_csv(data,sep=",")
        #df = df.iloc[3:]
        print(df) 
        """  df.columns=['Template',"LaunchProfileId",
        "LaunchPlanName","ChangeDate",
        "Version","UpdatedBy",
        "Origin_Location","Destination_Country",
        "Customer","Channel",
        "Other","Date_Type",
        "Target_Date","Qty",
        "Fulfillment_Scenario","Model_Type",
        "NodeModeOne","NodeModeTwo",
        "NodeModeThree","NodeModeFour",
        "NodeModeFive","NodeModeSix"] 
        """
        print(df)
        insert = "INSERT INTO [launchmodeldev].[dbo].[FactLaunchPlans](FileName,LaunchPlanId,LaunchProfileId,LaunchPlanName,ChangeDate,Version,UpdatedBy,Origin_Location,Destination_Country,Customer,Channel,Other,Date_Type,Target_Date,Qty,Fulfillment_Scenario,Model_Type,NodeModeOne,NodeModeTwo,NodeModeThree,NodeModeFour,NodeModeFive,NodeModeSix) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
        with conn.cursor() as cursor:
            for row in df.itertuples(): #iterrows():#itertuples()
                print(row)
                params = (FileName,id,launchID,row.LaunchPlanName,str(datetime.now()),row.Version,row.UpdatedBy,row.Origin_Location,row.Destination_Country,row.Customer,row.Channel,row.Other,row.Date_Type,str(row.Target_Date),row.Qty,row.Fulfillment_Scenario,row.Model_Type,row.NodeModeOne,row.NodeModeTwo,row.NodeModeThree,row.NodeModeFour,row.NodeModeFive,row.NodeModeSix)
                cursor.execute(str(insert),params)    
        cursor.close()
        #print(df)
    return print("OK") #redirect("launchprofile")


@app.route("/uploadfile", methods=['POST'])
def uploadFiles():
    if request.method=="POST":
        f = request.files['fileupload']
        test = request.form['form-select']
        print(test)
        form = request.form
        for key in form.keys():
            for value in form.getlist(key):
                print(key,":",value)
        fstring = f.read()
        text_obj = fstring.decode('UTF-8')
        data = io.StringIO(text_obj)
        df = pd.read_csv(data,sep=",")
        #csv_dicts = [{k: v for k, v in row.items()} for row in csv.DictReader(fstring.splitlines(), skipinitialspace=True)]  
        print(df)
    return("Success")

    
    #return render_template("templates/launchplans.html")



#Post, add an item to a launch
@app.route('/launch/<string:name>/sku',methods=['POST'])
def create_sku_in_launch(name):
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    request_data = request.get_json()
    for launch in launches:
        if launch['launchName'] == name:
            new_sku = {'skuName':request_data['skuName'],'price':request_data['price']}
            launch['sku'].append(new_sku)
            return jsonify(new_sku)
    return jsonify({"Message":"That launch is not found!"})

#@app.route('/launch/<string:name>/file',methods=['POST'])
#def upload_launch_file(name):
    

#GET a specific Launch
@app.route('/launch/<string:name>')
def get_launch(name):
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    for launch in launches:
        if launch["launchName"] == name:
            return jsonify(launch)
    return jsonify({"message":"Launch Not Found"})

#GET all skus in a specific launch 
@app.route('/launch/<string:name>/sku')
def get_sku_in_launch(name):
    #if not session.get("user"):
    #    return redirect(url_for("login"))
    for launch in launches:
        if launch["launchName"] == name:
            return jsonify({'sku':launch['sku']})
    return jsonify({"message":"Sku Not Found"})

'''

@app.errorhandler(Exception) 
def basic_error(e): 
    # fetch some info about the user from the request object 
    user_ip = request.remote_addr 
    requested_path = request.path 
 
    print("User with IP %s tried to access endpoint: %s" % (user_ip , requested_path)) 
    print("An error occurred: " + str(e) )
    responseobject = "An error occurred: " + str(e) 
    return responseobject

app.jinja_env.globals.update(_build_auth_code_flow=_build_auth_code_flow)  # Used in template

if __name__ == '__main__':
    app.run(port=5000)
